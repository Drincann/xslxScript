import argparse
import os
import openpyxl
from config import config
import util


class MonthlyPoint:
    def __init__(self, id: int, name: str, failureTimes, position: tuple, notuploadedTimes=None, processTimeTimes=None, health=None) -> None:
        self.id = int(id) if id != None else id
        self.name = name if not isinstance(name, str) else name.strip()
        self.failureTimes = failureTimes if failureTimes else 0
        self.row = position[0]
        self.col = position[1]
        self.notuploadedTimes = int(
            notuploadedTimes if notuploadedTimes != None else 0)
        self.processTimeTimes = processTimeTimes if processTimeTimes != None else 0
        self.health = int(health if health != None else 100)


class DailyPoint:
    def __init__(self, id: int, name: str, failureTimes, position: tuple, uploaded=None, processTime=None, health=None) -> None:
        self.id = int(id) if id != None else id
        self.name = name if not isinstance(name, str) else name.strip()
        self.failureTimes = failureTimes if failureTimes else 0
        self.row = position[0]
        self.col = position[1]
        self.uploaded = False if uploaded == '否' else True
        self.processTime = processTime if processTime != None else 0
        self.health = health if health != None else 100


class MonthlyReport:
    def __init__(self, filename, sheetname) -> None:
        self.filename = filename
        self.sheetname = sheetname
        self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
        self.sheet = self.workbook[self.sheetname]

    def getPointMap(self) -> dict[str, MonthlyPoint]:
        pointMap = {}
        for row in range(config.MonthlyPointNameRow, self.sheet.max_row + 1):
            pointId = self.sheet.cell(
                row=row, column=config.MonthlyPointIdCol).value

            name = self.sheet.cell(
                row=row, column=config.MonthlyPointNameCol).value
            failureTimes = self.sheet.cell(
                row=row, column=config.MonthlyPointFailureTimesCol).value
            notuploadedTimes = self.sheet.cell(
                row=row, column=config.MonthlyPointNotUploadedCol).value
            processTimeTimes = self.sheet.cell(
                row=row, column=config.MonthPointProcessTimeCol).value
            health = self.sheet.cell(
                row=row, column=config.MonthlyPointHealthCol).value
            pointMap[pointId] = MonthlyPoint(
                id=pointId,
                name=name,
                failureTimes=failureTimes,
                notuploadedTimes=notuploadedTimes,
                processTimeTimes=processTimeTimes,
                position=(row, config.MonthlyPointNameCol),
                health=health
            )
        return pointMap

    def update(self, dailyReport: "DailyReport", saveTo):
        dailyReportPointMap = dailyReport.getPointMap()
        monthlyPointMap = self.getPointMap()

        for dailyPoint in dailyReportPointMap.values():
            try:
                if dailyPoint.id in monthlyPointMap:
                    monthlyPoint = monthlyPointMap[dailyPoint.id]
                    monthlyPoint.failureTimes += dailyPoint.failureTimes
                    monthlyPoint.notuploadedTimes += (
                        1 if not dailyPoint.uploaded else 0)
                    monthlyPoint.processTimeTimes += (
                        1 if dailyPoint.processTime > 3 else 0)
                    monthlyPoint.health = util.calHealth(failureTimes=monthlyPoint.failureTimes,
                                                         notUploadedTimes=monthlyPoint.notuploadedTimes,
                                                         processTimeTimes=monthlyPoint.processTimeTimes)

                    self.sheet.cell(
                        row=monthlyPoint.row, column=config.MonthlyPointFailureTimesCol).value = monthlyPoint.failureTimes
                    self.sheet.cell(
                        row=monthlyPoint.row, column=config.MonthlyPointNotUploadedCol).value = monthlyPoint.notuploadedTimes
                    self.sheet.cell(
                        row=monthlyPoint.row, column=config.MonthPointProcessTimeCol).value = monthlyPoint.processTimeTimes
                    self.sheet.cell(
                        row=monthlyPoint.row, column=config.MonthlyPointHealthCol).value = monthlyPoint.health
                    if saveTo:
                        self.workbook.save(saveTo)
                    else:
                        self.workbook.save(self.filename)
                    self.sheet = self.workbook[self.sheetname]
            except Exception as e:
                print(e)

        if saveTo:
            self.workbook.save(saveTo)
        else:
            self.workbook.save(self.filename)
        self.sheet = self.workbook[self.sheetname]


class DailyReport:
    def __init__(self, filename, sheetname) -> None:
        self.filename = filename
        self.sheetname = sheetname
        self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
        self.sheet = self.workbook[self.sheetname]

    def getPointMap(self) -> dict[str, DailyPoint]:
        pointMap = {}
        for row in range(config.DailyPointNameRow, self.sheet.max_row + 1):
            pointId = self.sheet.cell(
                row=row, column=config.DailyPointIdCol).value

            name = self.sheet.cell(
                row=row, column=config.DailyPointNameCol).value
            failureTimes = self.sheet.cell(
                row=row, column=config.DailyPointFailureTimesCol).value
            uploaded = self.sheet.cell(
                row=row, column=config.DailyPointNotUploadedCol).value
            processTime = self.sheet.cell(
                row=row, column=config.DailyPointProcessTimeCol).value
            health = self.sheet.cell(
                row=row, column=config.DailyPointHealthCol).value

            pointMap[pointId] = DailyPoint(
                id=pointId,
                name=name,
                failureTimes=failureTimes,
                position=(row, config.DailyPointNameCol),
                uploaded=uploaded,
                processTime=processTime,
                health=health
            )
        return pointMap

    def update(self, monthlyReport: "MonthlyReport", saveTo):
        monthlyReportPointMap = monthlyReport.getPointMap()
        dailyPointMap = self.getPointMap()

        for monthlyPoint in monthlyReportPointMap.values():
            if monthlyPoint.id in dailyPointMap:
                try:
                    self.sheet.cell(
                        row=dailyPointMap[monthlyPoint.id].row,
                        column=config.DailyPointHealthCol
                    ).value = monthlyPoint.health

                except Exception as e:
                    print(e)
        if saveTo:
            self.workbook.save(saveTo)
        else:
            self.workbook.save(self.filename)
        self.sheet = self.workbook[self.sheetname]


def getArgs():
    parser = argparse.ArgumentParser(
        description='Update daily report and monthly report.')
    parser.add_argument('-m', '--monthly',
                        help='filename,sheet', required=True)
    parser.add_argument('-d', '--daily', help='filename,sheet', required=True)

    return parser.parse_args()


def checkArgs(args):
    if not args.monthly or not args.daily:
        print('Please specify filename and sheetname.')
        exit(1)
    try:
        dailyFilename, dailySheet = map(
            lambda item: item.strip(), args.daily.split(','))
        monthlyFilename, monthlySheet = map(
            lambda item: item.strip(), args.monthly.split(','))

    except Exception as e:
        print('Filename and sheetname should be separated by comma.')
        exit(1)

    if not os.path.exists(dailyFilename):
        print('Daily report file not exist.')
        exit(1)
    if not os.path.exists(monthlyFilename):
        print('Monthly report file not exist.')
        exit(1)

    return dailyFilename, dailySheet, monthlyFilename, monthlySheet


def main():
    # args
    dailyWorkbookFilename, dailyWorkbookSheetName,\
        monthlyWorkbookFilename, monthlyWorkbookSheetName\
        = checkArgs(getArgs())

    # monthlyReport
    monthlyReport = MonthlyReport(
        filename=monthlyWorkbookFilename,
        sheetname=monthlyWorkbookSheetName
    )
    dailyReport = DailyReport(
        filename=dailyWorkbookFilename,
        sheetname=dailyWorkbookSheetName
    )
    monthlyReport.update(
        dailyReport, f'./{util.basename(monthlyWorkbookFilename)}_updated{util.extname(monthlyWorkbookFilename)}')
    dailyReport.update(
        monthlyReport, f'./{util.basename(dailyWorkbookFilename)}_updated{util.extname(dailyWorkbookFilename)}')


main()
